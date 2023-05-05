from django.shortcuts import render,redirect
import openpyxl
from django.http import HttpResponse
from .models import student
from sign_in.models import student_sign_in
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
# Create your views here.


#这里写入导入班级学生excel数据的代码
@csrf_exempt
def excel_upload(request):
    if request.method == 'POST' and request.POST.get('password') == '0529':
        student_sign_in.objects.all().delete()#删除签到数据
        student.objects.all().delete()#上传之前把之前的班级数据删除
        file = request.FILES.get('file')
        workbook = openpyxl.load_workbook(file)
        worksheet = workbook.active
        rows = list(worksheet.iter_rows())
        headers = [cell.value for cell in rows[0]]
        # print(headers)

        objects = []
        #这种方法如果重复导入会报错，id重复了。可以去掉添加id
        for row in rows[1:]:
            data = {headers[i]: cell.value for i, cell in enumerate(row)}
            # print(data)
            # obj = MyModel(**data)
            objects.append(student(**data))
            # print(objects)
        # # 批量插入到数据库中
        student.objects.bulk_create(objects)
        return JsonResponse({'message': '上传成功'})

    return render(request, 'upload.html')

def delete_class_excel(request):
    student.objects.all().delete()
    return HttpResponse('删除成功')

#这里写入导入班级学生excel数据的代码
@csrf_exempt
def excel_upload_info(request):
    if request.method == 'POST':

        file = request.FILES.get('file')
        workbook = openpyxl.load_workbook(file)
        worksheet = workbook.active
        rows = list(worksheet.iter_rows())
        headers = [cell.value for cell in rows[0]]
        # print(headers)

        objects = []
        #这种方法如果重复导入会报错，id重复了。可以去掉添加id
        for row in rows[1:]:
            data = {headers[i]: cell.value for i, cell in enumerate(row)}
            # print(data)
            # obj = MyModel(**data)
            objects.append(student(**data))
            # print(objects)
        # # 批量插入到数据库中
        student.objects.bulk_create(objects)
        return HttpResponse('读取成功')
    return render(request, 'upload.html')